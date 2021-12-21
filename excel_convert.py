from openpyxl import load_workbook
import string

def convertExcel():
    workbook = load_workbook(filename="excel/music_rarity.xlsx", data_only=True)
    sheet = workbook.active
    total = []
    alphabet = list(string.ascii_uppercase)

    min = 2
    max = 4
    indexRow = 0
    for i in range(4):
        for row in sheet.iter_rows(min_row=3, max_row=40, min_col=min, max_col=max):
            if indexRow == 0:
                total.append({'trait_path': row[1].value, 'values': [], 'filename': [], 'weights': []})

            if indexRow == 1:
                total[i]['name'] = row[0].value

            if indexRow > 2:
                if row[0].value:
                    total[i]['values'].append(row[0].value)
                    total[i]['filename'].append(row[1].value)
                    total[i]['weights'].append(row[2].value)
            indexRow = indexRow + 1
        min = max + 2
        max = min + 2
        indexRow = 0
    print(total)
    return total